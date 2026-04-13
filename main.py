# app.py
# Production Version - BTU PDF Tabulation -> Single Row Excel
# ------------------------------------------------------------
# Install:
# uv venv
# source .venv/bin/activate
# uv pip install streamlit pdfplumber pandas openpyxl xlsxwriter

import streamlit as st
import pdfplumber
import pandas as pd
import re
import io
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="BTU PDF to Excel", layout="wide")

# ============================================================
# CONFIG
# ============================================================

SUB_COLS = ["ETE", "IA", "TOT", "LG", "GP", "EC", "CP"]

TOP_CROP = 25
BOTTOM_CROP = 140
LEFT_CROP = 5
RIGHT_CROP = 5

# ============================================================
# HELPERS
# ============================================================

def clean(x):
    if x is None:
        return ""
    return str(x).replace("\xa0", " ").strip()

def is_serial(x):
    return bool(re.fullmatch(r"\d+", clean(x)))

def is_roll(x):
    return bool(re.fullmatch(r"\d{2}[A-Z0-9]+", clean(x)))

def is_subject_code(x):
    x = clean(x)
    return bool(re.fullmatch(r"\d[A-Z]{2,3}\d*-\d+(?:\.\d+)?", x))

def normalize_spaces(txt):
    return re.sub(r"\s+", " ", clean(txt)).strip()

def get_lines(txt):
    txt = clean(txt)
    return [normalize_spaces(i) for i in txt.split("\n") if normalize_spaces(i)]

def safe_get(arr, idx):
    return arr[idx] if idx < len(arr) else ""

# ============================================================
# PDF EXTRACTION
# ============================================================

def extract_rows(uploaded_file):
    rows = []

    with pdfplumber.open(uploaded_file) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):

            width = page.width
            height = page.height

            crop = page.crop(
                (
                    LEFT_CROP,
                    TOP_CROP,
                    width - RIGHT_CROP,
                    height - BOTTOM_CROP
                )
            )

            tables = crop.extract_tables(
                table_settings={
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "intersection_tolerance": 5
                }
            )

            for table in tables:
                for row in table:
                    row = [clean(c) for c in row]
                    if any(cell != "" for cell in row):
                        rows.append(row)

    return rows

# ============================================================
# SUBJECT DETECTION
# ============================================================

def detect_subjects(rows):
    """
    Detect subject code + subject name from any row.
    Works on fixed BTU layout.
    """
    found = OrderedDict()

    for row in rows:
        for i in range(len(row)-1):

            code = clean(row[i])
            name = clean(row[i+1])

            if is_subject_code(code):
                if len(name) > 2:
                    key = (code, name)
                    found[key] = {
                        "code": code,
                        "name": name
                    }

    return list(found.values())

# ============================================================
# FIND STUDENT BLOCKS
# ============================================================

def find_student_start_indexes(rows):
    idxs = []

    for i, row in enumerate(rows):
        if len(row) >= 2:
            if is_serial(row[0]):
                idxs.append(i)

    return idxs

# ============================================================
# STUDENT IDENTITY PARSER
# ============================================================

def parse_identity(block_row):
    """
    Column 1 contains:
    Roll
    Enrollment
    Student
    Father
    Mother
    """

    roll_block = safe_get(block_row, 1)

    lines = get_lines(roll_block)

    roll = safe_get(lines, 0)
    enroll = safe_get(lines, 1)
    name = safe_get(lines, 2)
    father = safe_get(lines, 3)
    mother = safe_get(lines, 4)

    return {
        "Roll Number": roll,
        "Enrollment No.": enroll,
        "Name": name,
        "Father Name": father,
        "Mother Name": mother
    }

# ============================================================
# EXTRACT SGPA / CGPA
# ============================================================

def extract_sgpa_cgpa(block_rows):
    sgpa = ""
    cgpa = ""

    joined = " | ".join([" | ".join(r) for r in block_rows])

    m1 = re.search(r"SGPA[: ]+([0-9.]+)", joined, re.I)
    m2 = re.search(r"CGPA[: ]+([0-9.]+)", joined, re.I)

    if m1:
        sgpa = m1.group(1)

    if m2:
        cgpa = m2.group(1)

    return sgpa, cgpa

# ============================================================
# SUBJECT MARKS EXTRACTION
# ============================================================

def extract_subject_data(row):
    """
    BTU fixed columns
    """

    row = [clean(x) for x in row]

    vals = []

    indexes = [8, 9, 10, 11, 12, 13, 14]

    for idx in indexes:
        vals.append(row[idx] if idx < len(row) else "")

    return vals

# ============================================================
# MAIN PARSER
# ============================================================

def parse_students(rows, subjects):

    starts = find_student_start_indexes(rows)
    students = []

    for n, start in enumerate(starts):

        end = starts[n+1] if n+1 < len(starts) else len(rows)
        block = rows[start:end]

        if not block:
            continue

        first = block[0]

        student = {}
        student["S.NO"] = clean(first[0])

        # identity
        student.update(parse_identity(first))

        # subject rows
        subject_rows = []

        for r in block:
            for c in r:
                if is_subject_code(c):
                    subject_rows.append(r)
                    break

        for s_idx, sub in enumerate(subjects):

            prefix = f'{sub["name"]} ({sub["code"]})'

            row_match = None

            for r in subject_rows:
                if sub["code"] in r:
                    row_match = r
                    break

            if row_match:
                vals = extract_subject_data(row_match)
            else:
                vals = [""] * 7

            for col_name, val in zip(SUB_COLS, vals):
                student[f"{prefix}|{col_name}"] = val

        sgpa, cgpa = extract_sgpa_cgpa(block)

        student["SGPA"] = sgpa
        student["CGPA"] = cgpa

        students.append(student)

    return students

# ============================================================
# EXCEL EXPORT
# ============================================================

def build_excel(subjects, students):

    wb = Workbook()
    ws = wb.active
    ws.title = "Result"

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    fixed = [
        "S.NO",
        "Roll Number",
        "Enrollment No.",
        "Name",
        "Father Name",
        "Mother Name"
    ]

    col = 1

    # Fixed cols
    for item in fixed:
        ws.cell(row=1, column=col, value=item)
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1

    # Subjects
    for sub in subjects:

        title = f'{sub["name"]} ({sub["code"]})'

        start = col
        end = col + 6

        ws.cell(row=1, column=start, value=title)
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

        for i, sc in enumerate(SUB_COLS):
            ws.cell(row=2, column=start+i, value=sc)

        col += 7

    # SGPA / CGPA
    for item in ["SGPA", "CGPA"]:
        ws.cell(row=1, column=col, value=item)
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1

    # Style header
    for r in [1,2]:
        for c in range(1, col):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin

    # Rows
    row_no = 3

    for stu in students:

        c = 1

        for f in fixed:
            ws.cell(row=row_no, column=c, value=stu.get(f, ""))
            c += 1

        for sub in subjects:
            prefix = f'{sub["name"]} ({sub["code"]})'

            for sc in SUB_COLS:
                ws.cell(row=row_no, column=c, value=stu.get(f"{prefix}|{sc}", ""))
                c += 1

        ws.cell(row=row_no, column=c, value=stu.get("SGPA", ""))
        c += 1
        ws.cell(row=row_no, column=c, value=stu.get("CGPA", ""))

        row_no += 1

    # width
    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 30)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return out

# ============================================================
# UI
# ============================================================

st.title("📘 BTU PDF Result to Excel (Production Version)")

file = st.file_uploader("Upload BTU Result PDF", type=["pdf"])

if file:

    with st.spinner("Extracting rows from PDF..."):
        rows = extract_rows(file)

    st.success(f"Rows Extracted: {len(rows)}")

    with st.spinner("Detecting subjects..."):
        subjects = detect_subjects(rows)

    st.subheader("Detected Subjects")
    st.dataframe(pd.DataFrame(subjects))

    with st.spinner("Parsing students..."):
        students = parse_students(rows, subjects)

    st.success(f"Students Parsed: {len(students)}")

    if students:
        df = pd.DataFrame(students)
        st.dataframe(df)

        excel = build_excel(subjects, students)

        st.download_button(
            "📥 Download Excel",
            data=excel,
            file_name="BTU_Result_Output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )